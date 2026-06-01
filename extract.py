"""
BF6 Team Balancer - Excel → JSON data extraction script

Reads Excel file (3 columns: NAME, KD, KPM),
applies KPM offset correction, outputs players.json.
"""

import json
import sys
import os
from typing import Any, Optional
from openpyxl import load_workbook

# Default offset coefficient
DEFAULT_KPM_OFFSET = 1.313


def extract_players(xlsx_path: str, kpm_offset: float = DEFAULT_KPM_OFFSET) -> list:
    """Read player data from Excel, return a list of player dicts."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    players = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not row or row[0] is None:
            continue

        name = str(row[0]).strip()
        if not name:
            continue

        # Parse KD
        kd_raw = row[1] if len(row) > 1 else None
        kd = _parse_float(kd_raw)

        # Parse KPM
        kpm_raw = row[2] if len(row) > 2 else None
        kpm = _parse_float(kpm_raw)

        # Calculate adjusted KPM
        kpm_adjusted = round(kpm * kpm_offset, 2) if kpm is not None else None

        players.append({
            "name": name,
            "kd_raw": kd,
            "kpm_raw": kpm,
            "kd": kd,                    # KD needs no offset
            "kpm_adjusted": kpm_adjusted, # KPM after offset
        })

    wb.close()
    return players


def _parse_float(val: Any) -> Optional[float]:
    """Safely parse a float value, handling commas, percent signs, empty values, etc."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    s = str(val).strip()
    if not s:
        return None
    # Handle comma decimal points (European format)
    s = s.replace(",", ".")
    # Remove non-numeric characters (keep decimal point and minus sign)
    s = "".join(c for c in s if c.isdigit() or c == "." or c == "-")
    # Keep only the first decimal point (handle "1.2.3" -> "1.23")
    if s.count(".") > 1:
        first_dot = s.index(".")
        s = s[:first_dot + 1] + s[first_dot + 1:].replace(".", "")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def save_json(players: list, output_path: str) -> None:
    """Write player list to a JSON file."""
    data = {
        "meta": {
            "total_players": len(players),
            "kpm_offset": DEFAULT_KPM_OFFSET,
        },
        "players": players,
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已导出 {len(players)} 名玩家 → {output_path}")


def main():
    if len(sys.argv) < 2:
        print("用法: python extract.py <Excel文件路径> [KPM偏移系数]")
        print(f"示例: python extract.py players.xlsx 1.313")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"错误: 文件不存在 → {xlsx_path}")
        sys.exit(1)

    kpm_offset = DEFAULT_KPM_OFFSET
    if len(sys.argv) >= 3:
        try:
            kpm_offset = float(sys.argv[2])
        except ValueError:
            print(f"错误: 偏移系数必须是数字 → {sys.argv[2]}")
            sys.exit(1)

    players = extract_players(xlsx_path, kpm_offset)

    if not players:
        print("错误: Excel中没有找到玩家数据")
        sys.exit(1)

    # Output to players.json in the same directory as the Excel file
    output_path = os.path.join(os.path.dirname(os.path.abspath(xlsx_path)), "players.json")
    save_json(players, output_path)


if __name__ == "__main__":
    main()
